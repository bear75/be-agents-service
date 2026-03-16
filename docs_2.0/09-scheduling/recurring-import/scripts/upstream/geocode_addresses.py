#!/usr/bin/env python3
"""
Geocode addresses in Huddinge_Final.xlsx
Run this script on your local computer (requires internet access)

Usage:
    pip install pandas openpyxl requests
    python geocode_addresses.py Huddinge_Final.xlsx
"""

import pandas as pd
import requests
import time
import sys
import re
from openpyxl import load_workbook


def strip_address_for_geocode(street: str) -> str:
    """
    Remove apartment/floor/name details from street - geocoding works better with street only.
    Strips: LGH XX, Våning X, X trappor, trailing person names (e.g. MOSHE L EISHO).
    """
    if pd.isna(street) or not str(street).strip():
        return str(street) if not pd.isna(street) else ""
    s = str(street).strip()
    # LGH 1606, LGH XX, etc.
    s = re.sub(r"\s*LGH\s+\S+", "", s, flags=re.IGNORECASE)
    # Våning 2, våning 5, etc. (floor number)
    s = re.sub(r",?\s*våning\s+\d+", "", s, flags=re.IGNORECASE)
    # 3 trappor, etc. (X stairs)
    s = re.sub(r"\s+\d+\s+trappor", "", s, flags=re.IGNORECASE)
    # Trailing person names: 2+ words after street number (e.g. MOSHE L EISHO)
    s = re.sub(r"\s+[A-Za-zÀ-ÿ]{2,}(\s+[A-Za-zÀ-ÿ])?\s*[A-Za-zÀ-ÿ]{2,}$", "", s)
    # Clean trailing commas/whitespace
    return re.sub(r"[,;\s]+$", "", s).strip()


def normalize_postnr(postnr) -> str:
    """Remove spaces from postal code: 141 73 -> 14173."""
    if pd.isna(postnr):
        return ""
    return re.sub(r"\s+", "", str(postnr).strip())


def geocode_nominatim(address):
    """Geocode using Nominatim (free, 1 req/sec limit)"""
    try:
        url = "https://nominatim.openstreetmap.org/search"
        params = {"q": address, "format": "json", "limit": 1}
        headers = {"User-Agent": "HuddingeVisitGeocoder/1.0"}
        r = requests.get(url, params=params, headers=headers, timeout=10)
        data = r.json()
        if data:
            return float(data[0]['lat']), float(data[0]['lon'])
    except Exception as e:
        print(f"  Error: {e}")
    return None, None

def main(excel_file):
    print(f"Loading {excel_file}...")
    df = pd.read_excel(excel_file, sheet_name='Data')

    # Geocode office (once)
    print("\nGeocoding office...")
    office_raw = df['Kontor'].iloc[0] if 'Kontor' in df.columns else None
    if office_raw and ',' in str(office_raw):
        parts = str(office_raw).split(',')
        if len(parts) == 2:
            office_addr = f"{parts[1].strip()}, {parts[0].strip()}, Sweden"
            office_lat, office_lon = geocode_nominatim(office_addr)
            df['office_lat'] = office_lat
            df['office_lon'] = office_lon
            print(f"  Office: {office_lat}, {office_lon}")
            time.sleep(1)

    # Get unique client addresses
    print("\nGeocoding client addresses...")
    unique_addrs = df.groupby(['Kundnr', 'Gata', 'Postnr', 'Ort']).size().reset_index()[['Kundnr', 'Gata', 'Postnr', 'Ort']]

    # Build geocode lookup
    addr_coords = {}
    total = len(unique_addrs)
    success = 0

    for i, (_, row) in enumerate(unique_addrs.iterrows()):
        street_for_geocode = strip_address_for_geocode(row["Gata"])
        postnr = normalize_postnr(row["Postnr"])
        ort = str(row["Ort"]).strip()
        addr_full = (
            f"{street_for_geocode}, {postnr} {ort}, Sweden"
            if street_for_geocode
            else f"{postnr} {ort}, Sweden"
        )
        addr_fallback = f"{postnr} {ort}, Sweden"
        key = (row["Kundnr"], row["Gata"], row["Postnr"], row["Ort"])

        print(f"  [{i+1}/{total}] {row['Kundnr']}: {addr_full[:55]}...")
        lat, lon = geocode_nominatim(addr_full)
        used_fallback = False
        if not lat and street_for_geocode and addr_fallback != addr_full:
            time.sleep(1)
            lat, lon = geocode_nominatim(addr_fallback)
            used_fallback = bool(lat)
        addr_coords[key] = (lat, lon)

        if lat:
            success += 1
            print(f"    -> {lat}, {lon}" + (" (fallback)" if used_fallback else ""))
        else:
            print(f"    -> FAILED")

        time.sleep(1)  # Rate limit

    # Apply coordinates to dataframe
    def get_lat(row):
        key = (row['Kundnr'], row['Gata'], row['Postnr'], row['Ort'])
        return addr_coords.get(key, (None, None))[0]

    def get_lon(row):
        key = (row['Kundnr'], row['Gata'], row['Postnr'], row['Ort'])
        return addr_coords.get(key, (None, None))[1]

    df['client_lat'] = df.apply(get_lat, axis=1)
    df['client_lon'] = df.apply(get_lon, axis=1)

    print(f"\nGeocoded {success}/{total} addresses")

    # Save back to Excel
    output_file = excel_file.replace('.xlsx', '_geocoded.xlsx')
    print(f"\nSaving to {output_file}...")

    # Load workbook to preserve other sheets
    wb = load_workbook(excel_file)

    # Update Data sheet
    ws = wb['Data']

    # Find or add lat/lon columns
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

    # Add columns if they don't exist
    for col_name in ['office_lat', 'office_lon', 'client_lat', 'client_lon']:
        if col_name not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col_name)
            headers.append(col_name)

    # Write data
    for col_name in ['office_lat', 'office_lon', 'client_lat', 'client_lon']:
        col_idx = headers.index(col_name) + 1
        for row_idx, value in enumerate(df[col_name], start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_file)
    print(f"Done! Saved to {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python geocode_addresses.py <excel_file>")
        sys.exit(1)
    main(sys.argv[1])
